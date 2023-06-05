import os

from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


class CsvFile(object):
    def __init__(self, filename, report_details):
        self.dir_path = os.path.dirname(os.path.realpath(__file__))
        self.workbook = Workbook()
        self.filename = filename
        self.report_details = report_details

        self.create_title_tab()
        self.create_overview_tab()

    def create_title_tab(self):
        sheet = self.workbook.active
        sheet.title = "Title"
        sheet.merge_cells("A1:F36")
        merged_cell = sheet["A1"]
        merged_cell.fill = PatternFill(start_color="ebe4e2", end_color="ebe4e2", fill_type="solid")
        for i in range(48):
            sheet.row_dimensions[i+1].height = 15.75
        for col in ["A", "B", "C", "D", "E", "F"]:
            sheet.column_dimensions[col].width = 13
        # Add top left logo
        logo_path = os.path.join(self.dir_path, "report_assets", "logo1.png")
        img = Image(logo_path)
        img.width = 150
        img.height = 50
        sheet.add_image(img, "A1")
        # Add central image
        image_path = os.path.join(self.dir_path, "report_assets", "asset1.png")
        img = Image(image_path)
        img.anchor = "A6"
        img.width = 540
        img.height = 300
        sheet.add_image(img)
        # Add bottom right asset
        asset_path = os.path.join(self.dir_path, "report_assets", "asset2.png")
        img = Image(asset_path)
        img.anchor = "E35"
        img.width = 160
        img.height = 35
        sheet.add_image(img)

    def create_overview_tab(self):
        # Sheet general configuration
        sheet = self.workbook.create_sheet(title="Overview")
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 45
        for i in range(8):
            sheet.row_dimensions[i+1].height = 30

        # Sheet "Report Summary" header
        sheet.merge_cells("A1:B1")
        merged_cell = sheet["A1"]
        merged_cell.alignment = Alignment(horizontal="center", vertical="center")
        merged_cell.font = Font(name="Helvetica", size=10, bold=True)
        sheet["A1"].fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
        sheet["A1"] = "Report Summary"

        # Content headers
        for i, cell_content in enumerate(
                ["Report Name", "Description", "Date Generated (dd/mm/yyyy)",
                 "Total controls included", "AWS Accounts in scope",
                 "Assessment report selection"]):
            sheet[f"A{i + 2}"] = cell_content
            sheet[f"A{i + 2}"].font = Font(name="Helvetica", size=10, bold=True)
            sheet[f"A{i + 2}"].alignment = Alignment(horizontal="left", vertical="center")

        # Actual Content
        sheet["B2"] = self.filename.replace(".xlsx", "")
        sheet["B3"] = f'AWS {self.report_details["compliance_name"]} assessment report'
        sheet["B4"] = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M:%S") + " UTC"
        sheet["B5"] = str(self.report_details["total_rules"]) + " rules"
        sheet["B6"] = str(self.report_details["total_accounts"]) + " accounts"
        sheet["B7"] = str(self.report_details["total_violations"]) + " resources"

        # Footer
        sheet.merge_cells("A8:B8")
        merged_cell = sheet["A8"]
        merged_cell.alignment = Alignment(horizontal="center", vertical="center")
        merged_cell.font = Font(name="Helvetica", size=10, bold=True, color="FFFFFF")
        sheet["A8"].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        sheet["A8"] = "Generated by Lightlytics"

    def create_new_rule_sheet(self, violated_rule, rule_number, ws_accounts):
        # Sheet general configuration
        sheet = self.workbook.create_sheet(title=violated_rule["name"][0:30])
        sheet.column_dimensions['A'].width = 28
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 10
        for i in range(4):
            sheet.row_dimensions[i+1].height = 30

        # Sheet "Rule Name" header
        sheet.merge_cells("A1:D1")
        merged_cell = sheet["A1"]
        merged_cell.alignment = Alignment(horizontal="center", vertical="center")
        merged_cell.font = Font(name="Helvetica", size=10, bold=True)
        sheet["A1"].fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
        sheet["A1"] = f"{rule_number}. {violated_rule['name']}"

        # Content headers
        for i, cell_content in enumerate(["Control name", "Assessment report summary"]):
            sheet[f"A{i + 2}"] = cell_content
            sheet[f"A{i + 2}"].font = Font(name="Helvetica", size=10, bold=True)
            sheet[f"A{i + 2}"].alignment = Alignment(horizontal="left", vertical="center")

        # Actual Content
        sheet["B2"] = violated_rule["name"]
        total_resources = sum([a[1]["total_resources"] for a in violated_rule["violated_resources"].items()])
        total_violations = sum([len(a[1]["resource_ids"]) for a in violated_rule["violated_resources"].items()])
        total_compliant = total_resources - total_violations
        sheet["B3"] = f"{total_resources} ({total_compliant} Compliant, {total_violations} Non-compliant)"

        # Resources Table
        for i, account in enumerate(ws_accounts):
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            sheet[f"A{i + 5}"] = account["display_name"]
            sheet[f"A{i + 5}"].border = border
            sheet[f"B{i + 5}"] = account["cloud_account_id"]
            sheet[f"B{i + 5}"].border = border
            try:
                account_violation_count = len(
                    violated_rule["violated_resources"][account["cloud_account_id"]]["resource_ids"])
                total_account_resources = \
                    violated_rule["violated_resources"][account["cloud_account_id"]]["total_resources"]
                sheet[f"C{i + 5}"] = f"{account_violation_count} violations out of {total_account_resources}"
                sheet[f"C{i + 5}"].border = border
                ll_url = self.report_details['ll_url']
                ws_id = self.report_details['ws_id']
                sheet[f"D{i + 5}"].hyperlink = \
                    f"{ll_url}/w/{ws_id}/rules/{violated_rule['id']}" \
                    f"?f%5Baccount_id%5D%5B0%5D={account['cloud_account_id']}"
                font = Font(underline="single", color="0000FF")
                sheet[f"D{i + 5}"].font = font
                sheet[f"D{i + 5}"].value = "Evidence"
                sheet[f"D{i + 5}"].border = border
            except KeyError:
                sheet[f"C{i + 5}"] = "Found 0 violations"
                sheet[f"C{i + 5}"].border = border
                sheet[f"D{i + 5}"].border = border

        # Footer
        row_number = len(ws_accounts) + 5
        sheet.merge_cells(f"A{row_number}:D{row_number}")
        merged_cell = sheet[f"A{row_number}"]
        merged_cell.alignment = Alignment(horizontal="center", vertical="center")
        merged_cell.font = Font(name="Helvetica", size=12, bold=True, color="FFFFFF")
        sheet[f"A{row_number}"].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        sheet[f"A{row_number}"] = "Generated by Lightlytics"
        sheet.row_dimensions[row_number].height = 30

    def save_csv(self):
        self.workbook.save(self.filename)
